import openpyxl
import os

input_file = r'd:\Work\Speedruns\Games\Persona-5\Notes\P5R PC Notes (Advanced).xlsx'

if not os.path.exists(input_file):
    print(f"Error: {input_file} not found")
else:
    print(f"Loading '{input_file}'...")
    wb = openpyxl.load_workbook(input_file, data_only=True)
    ws = wb.worksheets[0]
    
    phase4_row_idx = -1
    
    # Find "Phase 4"
    for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
        # Check first few columns for "Phase 4"
        found = False
        for cell_val in row[:5]: 
            if cell_val and "Phase 4" in str(cell_val):
                phase4_row_idx = i
                found = True
                break
        if found:
            break
            
    if phase4_row_idx != -1:
        print(f"Found 'Phase 4' at Row {phase4_row_idx}")
        
        # Check a few rows after Phase 4 where the boxes are
        # Based on screenshot, "If Joker goes first:" is likley a few rows down.
        start_check = phase4_row_idx
        end_check = phase4_row_idx + 15 
        
        for i in range(start_check, end_check):
            row_cells = ws[i]
            print(f"\n--- Row {i} ---")
            for j, cell in enumerate(row_cells[:5]): # Check first 5 cols
                val = cell.value
                border = cell.border
                
                # Print border info if distinct
                has_border = False
                b_info = []
                if border.left.style: b_info.append(f"L:{border.left.style}")
                if border.right.style: b_info.append(f"R:{border.right.style}")
                if border.top.style: b_info.append(f"T:{border.top.style}")
                if border.bottom.style: b_info.append(f"B:{border.bottom.style}")
                
                if b_info:
                    print(f"  Col {j+1} ('{val}'): {', '.join(b_info)}")
                elif val:
                     print(f"  Col {j+1} ('{val}'): No Border")

    else:
        print("'Phase 4' not found.")
