import openpyxl
import os

input_file = r'd:\Work\Speedruns\Games\Persona-5\Notes\P5R PC Notes (Advanced).xlsx'

if not os.path.exists(input_file):
    print(f"Error: {input_file} not found")
else:
    print(f"Inspecting colors in '{input_file}'...")
    wb = openpyxl.load_workbook(input_file, data_only=True)
    ws = wb.worksheets[0]
    
    # Check first 20 rows
    for i in range(1, 21): # openpyxl is 1-indexed for rows in iteration usually, or we use ws.iter_rows
        # Get first cell of the row to check background
        cell = ws.cell(row=i, column=1)
        fill = cell.fill
        fg = fill.fgColor
        
        print(f"Row {i}: Type={fg.type}, RGB={fg.rgb}, Theme={fg.theme}, Tint={fg.tint}")
